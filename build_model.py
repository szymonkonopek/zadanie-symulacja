#!/usr/bin/env python3
"""Authoring script for the age-cohort stochastic epidemic simulation workbook.

Output: "model epidemii (kohorty wiekowe).xlsx" alongside this file.

The simulation runs entirely as Excel formulas (RAND / NORMINV / BINOM.INV).
This script is a one-shot builder — re-run after editing parameters at the top.
"""

import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName


# ---------- Parameters (single source of truth) ----------

T_MAX = 30
A_MAX = 10

INITIAL_STATE = {
    1: (60, 20, 10, 10),
    2: (60, 30, 10, 20),
    3: (70, 10, 5, 10),
    4: (60, 10, 10, 5),
    5: (20, 20, 7, 3),
    6: (10, 5, 6, 4),
    7: (10, 0, 0, 3),
    8: (0, 0, 0, 0),
    9: (0, 0, 0, 0),
    10: (0, 0, 0, 0),
}

BIRTH_MU_24, BIRTH_SIG_24 = 0.15, 0.02
BIRTH_MU_5P, BIRTH_SIG_5P = 0.10, 0.01

DEATH_MU_13, DEATH_SIG_13 = 0.20, 0.05
DEATH_MU_45, DEATH_SIG_45 = 0.30, 0.07
DEATH_MU_6P, DEATH_SIG_6P = 0.50, 0.15


def _phi(x):
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def _h_nat(a):
    if a >= A_MAX:
        return 1.0
    num = _phi(a - 6) - _phi(a - 7)
    den = 1 - _phi(a - 7)
    return num / den if den > 1e-12 else 1.0


H_NAT = {a: _h_nat(a) for a in range(1, A_MAX + 1)}


# ---------- Helpers ----------

HDR_FONT = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="DDEEFF")
SECTION_FONT = Font(bold=True, size=12)


def colL(a):
    """Excel column letter for age a (a=1 -> 'B')."""
    return get_column_letter(a + 1)


def write_header(ws, row, col_idx, text):
    cell = ws.cell(row=row, column=col_idx, value=text)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL


def fill_t_column(ws):
    write_header(ws, 1, 1, "t \\ a")
    for a in range(1, A_MAX + 1):
        write_header(ws, 1, a + 1, f"a={a}")
    for t in range(T_MAX + 1):
        ws.cell(row=t + 2, column=1, value=t)


# ---------- Build workbook ----------

wb = Workbook()


# ===== Sheet: Założenia =====
ws_zal = wb.active
ws_zal.title = "Założenia"
ws_zal.column_dimensions["A"].width = 32
for c in "BCDEF":
    ws_zal.column_dimensions[c].width = 12

write_header(ws_zal, 1, 1, "Parametr")
write_header(ws_zal, 1, 2, "Wartość")
ws_zal.cell(row=2, column=1, value="T_max (liczba kroków)")
ws_zal.cell(row=2, column=2, value=T_MAX)
ws_zal.cell(row=3, column=1, value="A_max (kohorty wiekowe)")
ws_zal.cell(row=3, column=2, value=A_MAX)

ws_zal.cell(row=5, column=1, value="Stopa urodzeń ~ Normal(μ, σ)").font = SECTION_FONT
write_header(ws_zal, 6, 1, "Pasmo wieku")
write_header(ws_zal, 6, 2, "μ")
write_header(ws_zal, 6, 3, "σ")
ws_zal.cell(row=7, column=1, value="2–4 u.j.c.")
ws_zal.cell(row=7, column=2, value=BIRTH_MU_24)
ws_zal.cell(row=7, column=3, value=BIRTH_SIG_24)
ws_zal.cell(row=8, column=1, value="5+ u.j.c.")
ws_zal.cell(row=8, column=2, value=BIRTH_MU_5P)
ws_zal.cell(row=8, column=3, value=BIRTH_SIG_5P)

ws_zal.cell(row=10, column=1, value="Śmiertelność choroby ~ Normal(μ, σ)").font = SECTION_FONT
write_header(ws_zal, 11, 1, "Pasmo wieku")
write_header(ws_zal, 11, 2, "μ")
write_header(ws_zal, 11, 3, "σ")
ws_zal.cell(row=12, column=1, value="1–3 u.j.c.")
ws_zal.cell(row=12, column=2, value=DEATH_MU_13)
ws_zal.cell(row=12, column=3, value=DEATH_SIG_13)
ws_zal.cell(row=13, column=1, value="4–5 u.j.c.")
ws_zal.cell(row=13, column=2, value=DEATH_MU_45)
ws_zal.cell(row=13, column=3, value=DEATH_SIG_45)
ws_zal.cell(row=14, column=1, value="6+ u.j.c.")
ws_zal.cell(row=14, column=2, value=DEATH_MU_6P)
ws_zal.cell(row=14, column=3, value=DEATH_SIG_6P)

ws_zal.cell(row=16, column=1,
            value="Naturalna śmiertelność (hazard z N(6, 1))").font = SECTION_FONT
write_header(ws_zal, 17, 1, "wiek a")
write_header(ws_zal, 17, 2, "h_nat(a)")
for a in range(1, A_MAX + 1):
    ws_zal.cell(row=17 + a, column=1, value=a)
    ws_zal.cell(row=17 + a, column=2, value=round(H_NAT[a], 6))

INIT_TABLE_TOP = 30
ws_zal.cell(row=INIT_TABLE_TOP, column=1, value="Stan początkowy (t=0)").font = SECTION_FONT
write_header(ws_zal, INIT_TABLE_TOP + 1, 1, "wiek a")
write_header(ws_zal, INIT_TABLE_TOP + 1, 2, "z")
write_header(ws_zal, INIT_TABLE_TOP + 1, 3, "o")
write_header(ws_zal, INIT_TABLE_TOP + 1, 4, "ca")
write_header(ws_zal, INIT_TABLE_TOP + 1, 5, "cb")
for a in range(1, A_MAX + 1):
    z, o, ca, cb = INITIAL_STATE[a]
    r = INIT_TABLE_TOP + 1 + a
    ws_zal.cell(row=r, column=1, value=a)
    ws_zal.cell(row=r, column=2, value=z)
    ws_zal.cell(row=r, column=3, value=o)
    ws_zal.cell(row=r, column=4, value=ca)
    ws_zal.cell(row=r, column=5, value=cb)
sum_row = INIT_TABLE_TOP + 1 + A_MAX + 1
write_header(ws_zal, sum_row, 1, "Σ")
for c in range(2, 6):
    L = get_column_letter(c)
    ws_zal.cell(row=sum_row, column=c,
                value=f"=SUM({L}{INIT_TABLE_TOP+2}:{L}{INIT_TABLE_TOP+1+A_MAX})")
ws_zal.cell(row=sum_row + 1, column=1, value="Razem populacja")
ws_zal.cell(row=sum_row + 1, column=2,
            value=f"=SUM(B{INIT_TABLE_TOP+2}:E{INIT_TABLE_TOP+1+A_MAX})")
ws_zal.cell(row=sum_row + 1, column=2).font = Font(bold=True)

notes_top = sum_row + 4
ws_zal.cell(row=notes_top, column=1,
            value="Kolejność operacji w kroku t → t+1").font = SECTION_FONT
notes = [
    "1. Wyznacz pch(t) = (Σ ca + Σ cb) / Σ żywych  (stan początku kroku).",
    "2. Wylosuj nowe zarażenia per kohorta:  pca(a, t+1) ~ Binom(z(a, t), pch(t)).",
    "3. Przejście choroby:  ca(a) → cb;  cb(a) → {śmierć z prawd. m(a) | wyzdrowienie → o}.",
    "4. Wygaśnięcie odporności:  o(t) → z(t+1) (jeśli przeżyje hazard naturalny).",
    "5. Naturalna śmiertelność:  Binom(stan, h_nat(a)) na każdym przedziale (przy wieku DOCELOWYM a).",
    "6. Starzenie:  kohorta a → a+1.  Kohorta a > A_max wymiera (h_nat = 1).",
    "7. Urodzenia: kohorty a ≥ 2  →  Binom(N(a, t), r(a)).  Część chora ~ udział chorych w kohorcie.",
]
for i, line in enumerate(notes):
    ws_zal.cell(row=notes_top + 1 + i, column=1, value=line)

defs = {
    "T_max": "$B$2",
    "A_max": "$B$3",
    "mu_r_24": "$B$7",
    "sig_r_24": "$C$7",
    "mu_r_5p": "$B$8",
    "sig_r_5p": "$C$8",
    "mu_m_13": "$B$12",
    "sig_m_13": "$C$12",
    "mu_m_45": "$B$13",
    "sig_m_45": "$C$13",
    "mu_m_6p": "$B$14",
    "sig_m_6p": "$C$14",
    "h_nat_tbl": f"$A$18:$B${17 + A_MAX}",
}
for name, ref in defs.items():
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'Założenia'!{ref}")


# ===== Compartment sheets z / o / ca / cb =====

def make_compartment_sheet(title, comp_idx_in_zal):
    """comp_idx_in_zal: 2=z, 3=o, 4=ca, 5=cb (column index in Założenia init table)."""
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 5
    for a in range(1, A_MAX + 1):
        ws.column_dimensions[colL(a)].width = 7
    fill_t_column(ws)
    # Row 2 (t=0): seed from Założenia
    for a in range(1, A_MAX + 1):
        init_row = INIT_TABLE_TOP + 1 + a
        ws.cell(row=2, column=a + 1,
                value=f"='Założenia'!{get_column_letter(comp_idx_in_zal)}{init_row}")
    return ws


ws_z = make_compartment_sheet("z", 2)
ws_o = make_compartment_sheet("o", 3)
ws_ca = make_compartment_sheet("ca", 4)
ws_cb = make_compartment_sheet("cb", 5)


# ===== Auxiliary sheets =====

def make_aux_sheet(title):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 5
    for a in range(1, A_MAX + 1):
        ws.column_dimensions[colL(a)].width = 9
    fill_t_column(ws)
    return ws


ws_pch = wb.create_sheet("pch")
ws_pch.column_dimensions["A"].width = 5
ws_pch.column_dimensions["B"].width = 12
write_header(ws_pch, 1, 1, "t")
write_header(ws_pch, 1, 2, "pch(t)")
for t in range(T_MAX + 1):
    ws_pch.cell(row=t + 2, column=1, value=t)

ws_r = make_aux_sheet("r_st")     # birth rate draws
ws_m = make_aux_sheet("m_st")     # mortality rate draws
ws_pca = make_aux_sheet("pca")    # new infections counts
ws_pu = make_aux_sheet("pu")      # disease death counts
ws_urodz = make_aux_sheet("urodz")          # births
ws_urodz_ch = make_aux_sheet("urodz_ch")    # sick births
ws_natz = make_aux_sheet("nat_z")
ws_nato = make_aux_sheet("nat_o")
ws_natca = make_aux_sheet("nat_ca")
ws_natcb = make_aux_sheet("nat_cb")


# ---- Formula population ----

def cell(sheet, t, a):
    """Cell reference for tick t, age a on a sheet by name."""
    return f"{sheet}!{colL(a)}{t + 2}"


def quoted(sheet):
    return sheet  # no spaces in our sheet names


# pch(t) = (Σ ca + Σ cb) / Σ alive — for every t (rows 2..32)
for t in range(T_MAX + 1):
    r = t + 2
    formula = (
        f"=IFERROR(("
        f"SUM(ca!B{r}:{colL(A_MAX)}{r})+SUM(cb!B{r}:{colL(A_MAX)}{r})"
        f")/("
        f"SUM(z!B{r}:{colL(A_MAX)}{r})+SUM(o!B{r}:{colL(A_MAX)}{r})"
        f"+SUM(ca!B{r}:{colL(A_MAX)}{r})+SUM(cb!B{r}:{colL(A_MAX)}{r})"
        f"),0)"
    )
    ws_pch.cell(row=r, column=2, value=formula)


# r_st(a, t) — birth rate draw per (age, tick)
# Cohorts < 2 produce 0.  Age band: 2-4 vs 5+.
for t in range(T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        if a < 2:
            formula = "=0"
        else:
            mu = "mu_r_24" if a <= 4 else "mu_r_5p"
            sig = "sig_r_24" if a <= 4 else "sig_r_5p"
            formula = f"=MAX(0,MIN(1,_xlfn.NORM.INV(RAND(),{mu},{sig})))"
        ws_r.cell(row=r, column=a + 1, value=formula)


# m_st(a, t) — disease mortality rate draw per (age, tick)
for t in range(T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        if a <= 3:
            mu, sig = "mu_m_13", "sig_m_13"
        elif a <= 5:
            mu, sig = "mu_m_45", "sig_m_45"
        else:
            mu, sig = "mu_m_6p", "sig_m_6p"
        formula = f"=MAX(0,MIN(1,_xlfn.NORM.INV(RAND(),{mu},{sig})))"
        ws_m.cell(row=r, column=a + 1, value=formula)


# pca(a, t) — new infections at destination tick t, source cohort age a.
# Drawn from z(a, t-1) with probability pch(t-1).
# Row 2 (t=0): 0.
for a in range(1, A_MAX + 1):
    ws_pca.cell(row=2, column=a + 1, value=0)
for t in range(1, T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        z_prev = cell("z", t - 1, a)
        pch_prev = f"pch!B{r-1}"
        formula = (
            f"=IFERROR(_xlfn.BINOM.INV(MAX(0,{z_prev}),"
            f"MAX(0,MIN(1,{pch_prev})),RAND()),0)"
        )
        ws_pca.cell(row=r, column=a + 1, value=formula)


# pu(a, t) — disease deaths during transition (t-1)->t, source cohort a.
# Drawn from cb(a, t-1) with probability m_st(a, t).
for a in range(1, A_MAX + 1):
    ws_pu.cell(row=2, column=a + 1, value=0)
for t in range(1, T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        cb_prev = cell("cb", t - 1, a)
        m_here = f"m_st!{colL(a)}{r}"
        formula = (
            f"=IFERROR(_xlfn.BINOM.INV(MAX(0,{cb_prev}),"
            f"MAX(0,MIN(1,{m_here})),RAND()),0)"
        )
        ws_pu.cell(row=r, column=a + 1, value=formula)


# urodz(a, t) — births from parent cohort a during transition (t-1)->t.
# N(a, t-1) parents x rate r_st(a, t).
for a in range(1, A_MAX + 1):
    ws_urodz.cell(row=2, column=a + 1, value=0)
for t in range(1, T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        if a < 2:
            ws_urodz.cell(row=r, column=a + 1, value=0)
            continue
        z_p = cell("z", t - 1, a)
        o_p = cell("o", t - 1, a)
        ca_p = cell("ca", t - 1, a)
        cb_p = cell("cb", t - 1, a)
        rate = f"r_st!{colL(a)}{r}"
        formula = (
            f"=IFERROR(_xlfn.BINOM.INV("
            f"MAX(0,{z_p}+{o_p}+{ca_p}+{cb_p}),"
            f"MAX(0,MIN(1,{rate})),RAND()),0)"
        )
        ws_urodz.cell(row=r, column=a + 1, value=formula)


# urodz_ch(a, t) — sick newborns from parent cohort a.
# P(sick | born) = (ca(a,t-1)+cb(a,t-1)) / N(a,t-1).
for a in range(1, A_MAX + 1):
    ws_urodz_ch.cell(row=2, column=a + 1, value=0)
for t in range(1, T_MAX + 1):
    r = t + 2
    for a in range(1, A_MAX + 1):
        if a < 2:
            ws_urodz_ch.cell(row=r, column=a + 1, value=0)
            continue
        z_p = cell("z", t - 1, a)
        o_p = cell("o", t - 1, a)
        ca_p = cell("ca", t - 1, a)
        cb_p = cell("cb", t - 1, a)
        births = f"urodz!{colL(a)}{r}"
        sick_share = (
            f"IFERROR(({ca_p}+{cb_p})/MAX(1,{z_p}+{o_p}+{ca_p}+{cb_p}),0)"
        )
        formula = (
            f"=IF({births}<=0,0,"
            f"IFERROR(_xlfn.BINOM.INV({births},MAX(0,MIN(1,{sick_share})),RAND()),0))"
        )
        ws_urodz_ch.cell(row=r, column=a + 1, value=formula)


# Helpers: raw post-disease-transition values (before natural death) ----------
# These are inlined inside nat_* and compartment formulas.
#
# For destination cohort a at tick t (a >= 2, t >= 1):
#   z_raw  = z(a-1, t-1) - pca(a-1, t) + o(a-1, t-1)
#   ca_raw = pca(a-1, t)
#   cb_raw = ca(a-1, t-1)
#   o_raw  = cb(a-1, t-1) - pu(a-1, t)
#
# For destination cohort a = 1 at tick t (t >= 1):
#   z_raw  = Σ (urodz(a, t) - urodz_ch(a, t)) for a in 2..A_MAX
#   ca_raw = Σ urodz_ch(a, t) for a in 2..A_MAX
#   cb_raw = 0
#   o_raw  = 0

def z_raw_expr(a, t):
    """Excel expression for z_raw at destination cohort a, tick t (t >= 1)."""
    if a == 1:
        r = t + 2
        end = colL(A_MAX)
        return (
            f"(SUM(urodz!C{r}:{end}{r})-SUM(urodz_ch!C{r}:{end}{r}))"
        )
    src = a - 1
    return (
        f"({cell('z', t - 1, src)}-{cell('pca', t, src)}"
        f"+{cell('o', t - 1, src)})"
    )


def ca_raw_expr(a, t):
    if a == 1:
        r = t + 2
        end = colL(A_MAX)
        return f"SUM(urodz_ch!C{r}:{end}{r})"
    src = a - 1
    return cell("pca", t, src)


def cb_raw_expr(a, t):
    if a == 1:
        return "0"
    src = a - 1
    return cell("ca", t - 1, src)


def o_raw_expr(a, t):
    if a == 1:
        return "0"
    src = a - 1
    return f"({cell('cb', t - 1, src)}-{cell('pu', t, src)})"


# nat_z, nat_o, nat_ca, nat_cb — natural deaths from each compartment at
# destination cohort (a, t), drawn from the raw post-disease pool.
def write_nat_sheet(ws_nat, raw_fn):
    for a in range(1, A_MAX + 1):
        ws_nat.cell(row=2, column=a + 1, value=0)
    for t in range(1, T_MAX + 1):
        r = t + 2
        for a in range(1, A_MAX + 1):
            raw = raw_fn(a, t)
            haz = f"VLOOKUP({a},h_nat_tbl,2,FALSE)"
            formula = (
                f"=IFERROR(_xlfn.BINOM.INV(MAX(0,{raw}),"
                f"MAX(0,MIN(1,{haz})),RAND()),0)"
            )
            ws_nat.cell(row=r, column=a + 1, value=formula)


write_nat_sheet(ws_natz, z_raw_expr)
write_nat_sheet(ws_nato, o_raw_expr)
write_nat_sheet(ws_natca, ca_raw_expr)
write_nat_sheet(ws_natcb, cb_raw_expr)


# Compartment final formulas:
#   compartment(a, t) = raw - nat
def write_compartment_formulas(ws, raw_fn, nat_sheet):
    for t in range(1, T_MAX + 1):
        r = t + 2
        for a in range(1, A_MAX + 1):
            raw = raw_fn(a, t)
            nat = f"{nat_sheet}!{colL(a)}{r}"
            formula = f"=MAX(0,{raw}-{nat})"
            ws.cell(row=r, column=a + 1, value=formula)


write_compartment_formulas(ws_z, z_raw_expr, "nat_z")
write_compartment_formulas(ws_o, o_raw_expr, "nat_o")
write_compartment_formulas(ws_ca, ca_raw_expr, "nat_ca")
write_compartment_formulas(ws_cb, cb_raw_expr, "nat_cb")


# ===== Wyniki sheet =====
ws_w = wb.create_sheet("Wyniki")
ws_w.column_dimensions["A"].width = 5
for c in "BCDEFGHIJK":
    ws_w.column_dimensions[c].width = 12
headers = [
    ("t", "A"),
    ("Σ z", "B"),
    ("Σ o", "C"),
    ("Σ ca", "D"),
    ("Σ cb", "E"),
    ("Σ zakaźnych", "F"),
    ("Σ żywych", "G"),
    ("Nowe zgony", "H"),
    ("Skumul. zgony", "I"),
    ("Σ urodzeń", "J"),
    ("Średni wiek", "K"),
]
for label, L in headers:
    write_header(ws_w, 1, ord(L) - 64, label)

for t in range(T_MAX + 1):
    r = t + 2
    end = colL(A_MAX)
    ws_w.cell(row=r, column=1, value=t)
    ws_w.cell(row=r, column=2, value=f"=SUM(z!B{r}:{end}{r})")
    ws_w.cell(row=r, column=3, value=f"=SUM(o!B{r}:{end}{r})")
    ws_w.cell(row=r, column=4, value=f"=SUM(ca!B{r}:{end}{r})")
    ws_w.cell(row=r, column=5, value=f"=SUM(cb!B{r}:{end}{r})")
    ws_w.cell(row=r, column=6, value=f"=D{r}+E{r}")
    ws_w.cell(row=r, column=7, value=f"=B{r}+C{r}+D{r}+E{r}")
    # New deaths this tick = Σ pu(a, t) + Σ nat_z + nat_o + nat_ca + nat_cb (all at row r)
    if t == 0:
        ws_w.cell(row=r, column=8, value=0)
    else:
        ws_w.cell(row=r, column=8, value=(
            f"=SUM(pu!B{r}:{end}{r})+SUM(nat_z!B{r}:{end}{r})"
            f"+SUM(nat_o!B{r}:{end}{r})+SUM(nat_ca!B{r}:{end}{r})"
            f"+SUM(nat_cb!B{r}:{end}{r})"
        ))
    if t == 0:
        ws_w.cell(row=r, column=9, value=0)
    else:
        ws_w.cell(row=r, column=9, value=f"=I{r-1}+H{r}")
    if t == 0:
        ws_w.cell(row=r, column=10, value=0)
    else:
        ws_w.cell(row=r, column=10, value=f"=SUM(urodz!B{r}:{end}{r})")
    # Mean age = Σ a*N(a) / Σ N(a)
    weighted = "+".join(
        f"{a}*(z!{colL(a)}{r}+o!{colL(a)}{r}+ca!{colL(a)}{r}+cb!{colL(a)}{r})"
        for a in range(1, A_MAX + 1)
    )
    ws_w.cell(row=r, column=11, value=f"=IFERROR(({weighted})/G{r},0)")


# Bottom: summary stats (single trajectory, current RAND state)
summary_top = T_MAX + 4
ws_w.cell(row=summary_top, column=1, value="Podsumowanie pojedynczej trajektorii").font = SECTION_FONT
labels = [
    ("Σ żywych w t=30",           f"=G{T_MAX+2}"),
    ("Skumul. zgony w t=30",      f"=I{T_MAX+2}"),
    ("Maks. zakaźnych",            f"=MAX(F2:F{T_MAX+2})"),
    ("Tick maks. zakaźności",      f"=MATCH(MAX(F2:F{T_MAX+2}),F2:F{T_MAX+2},0)-1"),
    ("Σ urodzeń całkowita",        f"=SUM(J2:J{T_MAX+2})"),
    ("Σ zgonów całkowita",         f"=I{T_MAX+2}"),
    ("Średni wiek w t=30",         f"=K{T_MAX+2}"),
    ("Wymarcie? (1=tak)",          f"=IF(G{T_MAX+2}=0,1,0)"),
]
for i, (lbl, formula) in enumerate(labels):
    ws_w.cell(row=summary_top + 1 + i, column=1, value=lbl)
    ws_w.cell(row=summary_top + 1 + i, column=2, value=formula)
    ws_w.cell(row=summary_top + 1 + i, column=1).font = Font(bold=True)


# Define named ranges to result aggregates for the Monte Carlo sheet
result_defs = {
    "res_zywe_t30":   f"$G${T_MAX+2}",
    "res_zgony_t30":  f"$I${T_MAX+2}",
    "res_peak":       f"$B${summary_top+3}",
    "res_peak_t":     f"$B${summary_top+4}",
    "res_urodz":      f"$B${summary_top+5}",
    "res_wiek_t30":   f"$K${T_MAX+2}",
    "res_wymarcie":   f"$B${summary_top+8}",
}
for name, ref in result_defs.items():
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'Wyniki'!{ref}")


# Chart on Wyniki
chart = LineChart()
chart.title = "Przebieg epidemii (pojedyncza trajektoria)"
chart.y_axis.title = "Liczba zwierząt"
chart.x_axis.title = "t (u.j.c.)"
chart.height = 11
chart.width = 22

data = Reference(ws_w, min_col=2, max_col=7, min_row=1, max_row=T_MAX + 2)
cats = Reference(ws_w, min_col=1, min_row=2, max_row=T_MAX + 2)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

deaths = Reference(ws_w, min_col=9, max_col=9, min_row=1, max_row=T_MAX + 2)
chart.add_data(deaths, titles_from_data=True)

ws_w.add_chart(chart, f"M2")


# ===== MonteCarlo sheet (stub with manual setup instructions) =====
ws_mc = wb.create_sheet("MonteCarlo")
ws_mc.column_dimensions["A"].width = 30
for c in "BCDEFGH":
    ws_mc.column_dimensions[c].width = 16

ws_mc.cell(row=1, column=1, value="Eksperyment Monte Carlo").font = SECTION_FONT
ws_mc.cell(row=2, column=1, value=(
    "Naciśnij F9 aby przeliczyć jedną trajektorię. "
    "Aby zebrać N replikacji, ustaw tabelę danych (Data Table) wg instrukcji poniżej."
))

ws_mc.cell(row=4, column=1, value="Instrukcja konfiguracji 1-wejściowej Data Table").font = SECTION_FONT
instr = [
    "1. Zaznacz zakres B10:H510.",
    "2. Wstążka: Dane → Analiza warunkowa → Tabela danych…",
    "   (Mac UI po angielsku: Data → What-If Analysis → Data Table…)",
    "3. „Komórka wejściowa wiersza” – ZOSTAW PUSTE.",
    "4. „Komórka wejściowa kolumny” – wskaż dowolną pustą komórkę, np. $A$1.",
    "5. OK. Excel uruchomi 500 przeliczeń – w kolumnach C..H pojawią się próbki,",
    "   a statystyki w wierszach 513–517 (średnia / odch.std / p5 / p50 / p95) policzą się same.",
    "6. (Opcjonalnie) Plik → Opcje → Formuły → „Automatyczne z wyjątkiem tabel danych”,",
    "   żeby tabela nie przeliczała się przy każdej zmianie komórki.",
    "",
    "Uwaga: do pierwszego uruchomienia Tabeli Danych komórki C11:H510 są puste,",
    "więc statystyki poniżej pokazują „—”. To normalne – znikną po uruchomieniu.",
]
for i, line in enumerate(instr):
    ws_mc.cell(row=5 + i, column=1, value=line)

# Data Table layout (1-input column-variable):
#   B10        | C10 | D10 | … | H10   ← corner blank + result formulas
#   1          |     |     |   |       ← inputs in col B, results filled by Excel
#   2          |     |     |   |
#   …          |     |     |   |
#   500        |     |     |   |
# Select B10:H510, run Data Table with column input = any unused cell.
ws_mc.cell(row=10, column=2, value=None)  # corner blank
mc_results = [
    ("Σ żywych t=30",      "=res_zywe_t30"),
    ("Skumul. zgony t=30", "=res_zgony_t30"),
    ("Maks. zakaźnych",    "=res_peak"),
    ("Tick maks.",         "=res_peak_t"),
    ("Σ urodzeń",          "=res_urodz"),
    ("Wymarcie?",          "=res_wymarcie"),
]
# Labels in row 9 (so user can see what each column means without colliding
# with the Data Table's formula row).
for j, (lbl, _formula) in enumerate(mc_results):
    col_idx = 3 + j
    ws_mc.cell(row=9, column=col_idx, value=lbl).font = HDR_FONT
    ws_mc.cell(row=9, column=col_idx).fill = HDR_FILL
# Formulas in row 10 (the Data Table formula row).
for j, (_lbl, formula) in enumerate(mc_results):
    col_idx = 3 + j
    ws_mc.cell(row=10, column=col_idx, value=formula)

# Rep index column (B11..B510)
for i in range(500):
    ws_mc.cell(row=11 + i, column=2, value=i + 1)

# Summary stats below
stats_top = 512
ws_mc.cell(row=stats_top, column=1, value="Statystyki (po uruchomieniu Data Table)").font = SECTION_FONT
stat_rows = [
    ("średnia",   "AVERAGE"),
    ("odch. std", "STDEV.S"),
    ("p5",         "PERCENTILE.INC"),
    ("p50",        "PERCENTILE.INC"),
    ("p95",        "PERCENTILE.INC"),
]
percentile_args = {"p5": 0.05, "p50": 0.5, "p95": 0.95}
for i, (lbl, fn) in enumerate(stat_rows):
    r = stats_top + 1 + i
    ws_mc.cell(row=r, column=1, value=lbl).font = Font(bold=True)
    for j in range(len(mc_results)):
        col_idx = 3 + j
        L = get_column_letter(col_idx)
        rng = f"{L}11:{L}510"
        # Excel 2010+ stats functions need _xlfn. prefix in OOXML storage.
        # Wrap in IFERROR + COUNT-guard so the cell shows "—" until the Data
        # Table has actually populated the range (otherwise AVERAGE/STDEV/PCT
        # of an empty range raise #DIV/0! and #NUM!).
        prefixed = f"_xlfn.{fn}" if fn in ("STDEV.S", "PERCENTILE.INC") else fn
        if fn == "PERCENTILE.INC":
            arg = percentile_args[lbl]
            inner = f"{prefixed}({rng},{arg})"
        else:
            inner = f"{prefixed}({rng})"
        formula = f'=IFERROR(IF(COUNT({rng})=0,"—",{inner}),"—")'
        ws_mc.cell(row=r, column=col_idx, value=formula)


# ===== Workbook-level calc properties =====
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True

# Move Założenia to position 0 (already there as wb.active), then z/o/ca/cb...
# Already in creation order. Ensure Wyniki & MonteCarlo come after aux sheets — they do.

out = "/Users/szymon/Desktop/zadanie-symulacja/model epidemii (kohorty wiekowe).xlsx"
wb.save(out)
print(f"Wrote {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"H_NAT table: {H_NAT}")
